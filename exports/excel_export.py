"""
Board-ready Excel report generator.

Produces a professionally formatted .xlsx file with 4 tabs:
  1. Executive Summary — key KPI metrics table
  2. Full GTM Financial Model — all 24 months of data
  3. Scenario Analysis — Bull / Base / Bear projections
  4. AI Recommendations — agent narrative and action items
"""

import io
import json
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from typing import Any

from models.financial_model import (
    compute_gtm_metrics,
    get_latest_metrics_summary,
    get_health_status,
    BENCHMARKS,
)


# ──────────────────────────────────────────────────────────────
# Style constants
# ──────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="CC785C", end_color="CC785C", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="16213E", end_color="16213E", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, color="E0E0E0", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color="1a1a2e", size=14)
BODY_FONT = Font(name="Calibri", size=10, color="333333")
HEALTHY_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")
CRITICAL_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
WRAP_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _auto_column_width(ws: Any, min_width: int = 12, max_width: int = 40) -> None:
    """Auto-fit column widths based on cell content.

    Args:
        ws: Openpyxl worksheet.
        min_width: Minimum column width.
        max_width: Maximum column width.
    """
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def _write_title_row(ws: Any, row: int, title: str, col_span: int = 6) -> None:
    """Write a styled title row across multiple columns.

    Args:
        ws: Openpyxl worksheet.
        row: Row number to write to.
        title: Title text.
        col_span: Number of columns to merge.
    """
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _build_executive_summary(
    wb: Workbook,
    df: pd.DataFrame,
    company_name: str,
    report_period: str,
) -> None:
    """Build the Executive Summary tab.

    Args:
        wb: Openpyxl workbook.
        df: Raw GTM dataset.
        company_name: Company name for the header.
        report_period: Reporting period string.
    """
    ws = wb.active
    ws.title = "Executive Summary"

    # Title
    _write_title_row(ws, 1, f"{company_name} — GTM Financial Executive Summary", 6)
    ws.cell(row=2, column=1, value=f"Report Period: {report_period}").font = Font(
        name="Calibri", italic=True, size=10, color="666666"
    )

    # KPI table
    row = 4
    headers = ["Metric", "Current Value", "Benchmark", "Status", "Trend"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    latest = get_latest_metrics_summary(df)
    metrics_df = compute_gtm_metrics(df)

    kpi_rows = [
        ("Annual Recurring Revenue (ARR)", f"${latest.get('arr', 0):,.0f}", "Growing MoM", "arr"),
        ("Monthly Recurring Revenue (MRR)", f"${latest.get('mrr', 0):,.0f}", "Growing MoM", "mrr"),
        ("MoM Growth Rate", f"{latest.get('mom_growth_pct', 0):.1f}%", ">5%", "mom_growth_pct"),
        ("YoY Growth Rate", f"{latest.get('yoy_growth_pct', 0):.1f}%", ">50%", "yoy_growth_pct"),
        ("Customer Acquisition Cost (CAC)", f"${latest.get('cac', 0):,.0f}", "<$500", "cac"),
        ("Lifetime Value (LTV)", f"${latest.get('ltv', 0):,.0f}", ">3× CAC", "ltv"),
        ("LTV:CAC Ratio", f"{latest.get('ltv_cac_ratio', 0):.1f}x", ">3.0x", "ltv_cac_ratio"),
        ("Payback Period", f"{latest.get('payback_months', 0):.1f} months", "<12 months", "payback_months"),
        ("Net Revenue Retention (NRR)", f"{latest.get('nrr_pct', 0):.1f}%", ">110%", "nrr_pct"),
        ("Burn Multiple", f"{latest.get('burn_multiple', 0):.1f}x", "<2.0x", "burn_multiple"),
        ("Sales Productivity", f"${latest.get('sales_productivity', 0):,.0f}", "Increasing", "sales_productivity"),
        ("Pipeline Coverage", f"{latest.get('pipeline_coverage', 0):.1f}x", ">3.0x", "pipeline_coverage"),
        ("Total Customers", f"{latest.get('cumulative_customers', 0):,.0f}", "Growing", "cumulative_customers"),
    ]

    for i, (label, value, benchmark, key) in enumerate(kpi_rows):
        r = row + 1 + i
        status = get_health_status(latest.get(key, 0), key) if key in BENCHMARKS else "neutral"

        # Determine trend
        if len(metrics_df) >= 2 and key in metrics_df.columns:
            prev = metrics_df[key].iloc[-2] if not pd.isna(metrics_df[key].iloc[-2]) else 0
            curr = latest.get(key, 0) or 0
            trend = "↑ Improving" if curr > prev else "↓ Declining" if curr < prev else "→ Flat"
        else:
            trend = "—"

        status_text = {"healthy": "✅ Healthy", "warning": "⚠️ Warning", "critical": "🔴 Critical"}.get(
            status, "ℹ️ N/A"
        )

        cells_data = [label, value, benchmark, status_text, trend]
        for col_idx, val in enumerate(cells_data, 1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN if col_idx > 1 else Alignment(horizontal="left")

            # Conditional fill
            if col_idx == 4:
                if status == "healthy":
                    cell.fill = HEALTHY_FILL
                elif status == "warning":
                    cell.fill = WARNING_FILL
                elif status == "critical":
                    cell.fill = CRITICAL_FILL

    _auto_column_width(ws)


def _build_full_model(wb: Workbook, df: pd.DataFrame) -> None:
    """Build the Full GTM Financial Model tab.

    Args:
        wb: Openpyxl workbook.
        df: Raw GTM dataset.
    """
    ws = wb.create_sheet("GTM Financial Model")

    metrics_df = compute_gtm_metrics(df)

    # Columns to include
    columns = [
        ("Date", "date"),
        ("MRR ($)", "mrr"),
        ("ARR ($)", "arr"),
        ("MoM Growth %", "mom_growth_pct"),
        ("YoY Growth %", "yoy_growth_pct"),
        ("CAC ($)", "cac"),
        ("LTV ($)", "ltv"),
        ("LTV:CAC", "ltv_cac_ratio"),
        ("Payback (mo)", "payback_months"),
        ("Sales Prod ($)", "sales_productivity"),
        ("NRR %", "nrr_pct"),
        ("Burn Multiple", "burn_multiple"),
        ("New Customers", "new_customers"),
        ("Churned", "churned_customers"),
        ("Total Customers", "cumulative_customers"),
    ]

    available = [(label, key) for label, key in columns if key in metrics_df.columns]

    # Headers
    for col_idx, (label, _) in enumerate(available, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # Data rows
    for row_idx, (_, row) in enumerate(metrics_df.iterrows(), 2):
        for col_idx, (_, key) in enumerate(available, 1):
            val = row.get(key)
            if isinstance(val, pd.Timestamp):
                val = val.strftime("%Y-%m")
            elif isinstance(val, (np.floating, float)) and pd.isna(val):
                val = "—"
            elif isinstance(val, (np.floating, float)):
                val = round(float(val), 2)
            elif isinstance(val, (np.integer,)):
                val = int(val)

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

            # Conditional formatting for key metrics
            if key in BENCHMARKS and isinstance(val, (int, float)):
                status = get_health_status(float(val), key)
                if status == "healthy":
                    cell.fill = HEALTHY_FILL
                elif status == "warning":
                    cell.fill = WARNING_FILL
                elif status == "critical":
                    cell.fill = CRITICAL_FILL

    _auto_column_width(ws)


def _build_scenario_tab(wb: Workbook, scenarios_json: str) -> None:
    """Build the Scenario Analysis tab.

    Args:
        wb: Openpyxl workbook.
        scenarios_json: JSON string with Bull/Base/Bear projections.
    """
    ws = wb.create_sheet("Scenario Analysis")

    scenarios = json.loads(scenarios_json) if scenarios_json and scenarios_json != "{}" else {}

    if not scenarios:
        ws.cell(row=1, column=1, value="No scenario data available. Run the AI agent first.").font = BODY_FONT
        return

    _write_title_row(ws, 1, "12-Month Scenario Projections — Bull / Base / Bear", 8)

    row = 3
    for key in ["bull", "base", "bear"]:
        if key not in scenarios:
            continue

        scenario = scenarios[key]
        label = scenario.get("label", key.title())

        # Scenario header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        row += 1

        # Assumptions
        assumptions = scenario.get("assumptions", {})
        ws.cell(row=row, column=1, value="Assumptions:").font = Font(bold=True, size=10)
        for i, (a_key, a_val) in enumerate(assumptions.items()):
            ws.cell(row=row, column=2 + i, value=f"{a_key}: {a_val}").font = BODY_FONT
        row += 1

        # Month-12 projection
        m12 = scenario.get("month_12_projection", {})
        headers = list(m12.keys())
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h.upper().replace("_", " "))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
        row += 1

        for col_idx, h in enumerate(headers, 1):
            val = m12[h]
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
        row += 1

        # Monthly trajectory
        trajectory = scenario.get("monthly_trajectory", [])
        if trajectory:
            traj_headers = ["Month", "MRR ($)", "ARR ($)"]
            for col_idx, h in enumerate(traj_headers, 1):
                cell = ws.cell(row=row, column=col_idx, value=h)
                cell.font = Font(bold=True, size=10)
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGN
            row += 1

            for t in trajectory:
                ws.cell(row=row, column=1, value=t.get("month", "")).font = BODY_FONT
                ws.cell(row=row, column=2, value=t.get("mrr", "")).font = BODY_FONT
                ws.cell(row=row, column=3, value=t.get("arr", "")).font = BODY_FONT
                for c in range(1, 4):
                    ws.cell(row=row, column=c).border = THIN_BORDER
                    ws.cell(row=row, column=c).alignment = CENTER_ALIGN
                row += 1

        row += 2  # spacing between scenarios

    _auto_column_width(ws)


def _build_recommendations_tab(
    wb: Workbook,
    narrative: str,
    recommendations_json: str,
) -> None:
    """Build the AI Recommendations tab.

    Args:
        wb: Openpyxl workbook.
        narrative: The agent's executive narrative memo.
        recommendations_json: JSON string with recommendations.
    """
    ws = wb.create_sheet("AI Recommendations")

    _write_title_row(ws, 1, "AI Strategy Agent — Executive Recommendations", 6)
    ws.cell(row=2, column=1, value="Generated by Claude Opus GTM Finance Intelligence Agent").font = Font(
        italic=True, size=10, color="666666"
    )

    row = 4

    # Write narrative
    if narrative:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value="EXECUTIVE MEMO")
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        row += 1

        # Split narrative into paragraphs
        paragraphs = narrative.split("\n")
        for para in paragraphs:
            para = para.strip()
            if para:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                cell = ws.cell(row=row, column=1, value=para)
                cell.font = BODY_FONT
                cell.alignment = WRAP_ALIGN
                ws.row_dimensions[row].height = max(15, len(para) // 5)
                row += 1
        row += 1

    # Write structured recommendations
    recs = json.loads(recommendations_json) if recommendations_json and recommendations_json != "{}" else {}
    rec_list = recs.get("recommendations", [])

    if rec_list:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value="STRATEGIC RECOMMENDATIONS")
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        row += 1

        headers = ["Priority", "Area", "Title", "Current State", "Actions", "Expected Impact"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
        row += 1

        for rec in rec_list:
            actions = "\n".join(f"• {a}" for a in rec.get("actions", []))
            data = [
                rec.get("priority", ""),
                rec.get("area", ""),
                rec.get("title", ""),
                rec.get("current_state", ""),
                actions,
                rec.get("expected_impact", ""),
            ]
            for col_idx, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                cell.alignment = WRAP_ALIGN
            ws.row_dimensions[row].height = 60
            row += 1

    _auto_column_width(ws, max_width=50)


def generate_board_report(
    df: pd.DataFrame,
    company_name: str = "SaaS Corp",
    report_period: str = "FY2024-2025",
    agent_narrative: str = "",
    scenarios_json: str = "{}",
    recommendations_json: str = "{}",
) -> io.BytesIO:
    """Generate a complete board-ready Excel report.

    Args:
        df: Raw GTM dataset.
        company_name: Company name for report headers.
        report_period: Reporting period string.
        agent_narrative: The AI agent's executive memo text.
        scenarios_json: JSON string with scenario projections.
        recommendations_json: JSON string with strategic recommendations.

    Returns:
        BytesIO buffer containing the .xlsx file.
    """
    wb = Workbook()

    _build_executive_summary(wb, df, company_name, report_period)
    _build_full_model(wb, df)
    _build_scenario_tab(wb, scenarios_json)
    _build_recommendations_tab(wb, agent_narrative, recommendations_json)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
